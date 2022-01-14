from email.policy import default
from sre_parse import State
from tkinter.tix import TCL_DONT_WAIT
import openpyxl
from datetime import datetime
import os
from flask import Flask, request, render_template, redirect
from flask.helpers import url_for
import openpyxl
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import requests
import json

from sqlalchemy.orm import defaultload

app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///test.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db = SQLAlchemy(app)


# class Data(db.Model):
#     id = db.Column(db.Integer, primary_key=True)
#     mobile_no = db.Column(db.String(10))
#     whatsappnumber = db.Column(db.Integer)
#     name = db.Column(db.String(1000))
#     description = db.Column(db.String(1000))
#     email = db.Column(db.String(1000))
#     created_date = db.Column(db.DateTime, default=datetime.now)
#     status = db.Column(db.Integer, default=0)
#     visible = db.Column(db.Integer, default=0)
#     deliverycount_indicator = db.Column(db.Integer, default=0)
#     hasdynamiclocation = db.Column(db.Integer, default=0)
#     serviceabledistance = db.Column(db.Integer)
#     rating = db.Column(db.Integer, nullable=False)
#     image = db.Column(db.String(1000))
#     multipleoutlet = db.Column(db.Integer, default=0)
#     multiple_service = db.Column(db.Integer, default=0)
#     popular_serviceprovider = db.Column(db.Integer, default=0)
#     discount = db.Column(db.Integer)
#     totalrating = db.Column(db.Float)
#     totalrating_user = db.Column(db.Float)
#     updated_datetime = db.Column(db.DateTime, default=datetime.now)
#     admin_id = db.Column(db.Integer)
#     is_delete = db.Column(db.Integer, default=0)
#     subscription_date = db.Column(db.DateTime)
#     is_paid = db.Column(db.Integer, default=0)
#     is_verified = db.Column(db.Integer, default=0)
#     website = db.Column(db.String(1000))
#     username = db.Column(db.String(1000))


class Address(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    pincode = db.Column(db.Integer, nullable=False)
    longitude = db.Column(db.Float, default=83.3612555)
    latitude = db.Column(db.Float, default=26.7828471)
    address_line_1 = db.Column(db.String(1000))
    address_line_2 = db.Column(db.String(1000))
    landmark = db.Column(db.String(1000))
    town = db.Column(db.String(1000))
    state = db.Column(db.String(1000))
    district = db.Column(db.String(1000))


@app.route("/", methods=["POST", "GET"])
def homepage():
    db.create_all()
    db.session.commit()
    if request.method == "POST":
        file = request.files['file']
        path = f"static/{file.filename}"
        file.save(path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        max_data = ws.max_row

        # Address Values
        def address_lines(value):
            if value == None:
                return None, None
            a = value.split(",", 1)
            address_line_1, address_line_2 = a[0], a[1].strip()
            print(address_line_1, address_line_2)
            return address_line_1, address_line_2

        # API Requests
        def location_request(pincode):
            url = "https://api.postalpincode.in/pincode/" + str(pincode)
            response = requests.get(url).json()
            value = response[0]['PostOffice'][0]
            # print(value)
            print("You are returning this:",
                  value['Block'], value['District'], value['State'])
            return value['Block'], value['District'], value['State']

        for i in (2, max_data+1):
            pincode = ws.cell(row=i, column=6).value
            # longitude
            # latitude
            address_line_1, address_line_2 = address_lines(
                ws.cell(row=i, column=5).value)
            # landmark
            town, district, state = location_request(pincode)
            print(town, district, state)
            address = Address(pincode=pincode, address_line_1=address_line_1,
                              address_line_2=address_line_2, town=town, district=district, state=state)
            db.session.add(address)
            db.session.flush()
            db.session.commit()

        return redirect("/")
    return render_template('homepage.html')

# @app.route("/", methods=["POST", "GET"])
# def homepage():
#     db.create_all()
#     db.session.commit()
#     if request.method == "POST":
#         file = request.files['file']
#         path = f"static/{file.filename}"
#         file.save(path)
#         wb = openpyxl.load_workbook(path)
#         ws = wb.active
#         max_data = ws.max_row
#         for i in range(2,max_data+1):
#             def phone_no(value):
#                 value = str(value)
#                 l = len(value)
#                 if l >= 10 and l<=13:
#                     if l == 10:
#                         new_val = "+91" + value
#                         return new_val
#                     if l == 13 and value[0:3] == "+91":
#                         new_val = value
#                         return new_val
#                     if l == 12 and value[0:2] == "91":
#                         new_val = "+" + value
#                         return new_val
#                     return None
#                 return None

#             def convert(i,j):
#                 value = ws.cell(column=j, row=i).value
#                 if value == "Yes":
#                     return 1
#                 return 0

#             checkforphone = phone_no(ws.cell(column=9,row=i).value)
#             if (checkforphone == None):
#                 print("Number not recognised")
#                 continue
#             else:
#                 print("Number recognised :", checkforphone)
#                 mobile_no = checkforphone
#             whatsappnumber = ws.cell(column=8, row=i).value
#             name = ws.cell(column=3, row=i).value
#             description = ws.cell(column=2, row=i).value
#             email = ws.cell(column=7, row=i).value
#             # created_date
#             # status
#             # visible
#             # deliverycount_indicator
#             # hasdynamiclocation
#             serviceabledistance = 0
#             rating = ws.cell(column=14, row=i).value
#             image = None
#             # multipleoutlet
#             # multiple_service
#             # popular_serviceprovider
#             discount = None
#             totalrating = ws.cell(column=15, row=i).value
#             totalrating_user = None
#             # updated_datetime
#             admin_id = None
#             # is_delete
#             subscription_date = None
#             is_paid = convert(i, 18)
#             is_verified = convert(i,17)
#             website = ws.cell(column=16, row=i).value
#             username = None
#             data=Data(mobile_no=mobile_no,whatsappnumber=whatsappnumber,name=name,description=description,email=email,serviceabledistance=serviceabledistance,rating=rating,image=image,discount=discount,totalrating=totalrating,totalrating_user=totalrating_user,admin_id=admin_id,subscription_date=subscription_date,is_paid=is_paid,is_verified=is_verified,website=website,username=username)
#             db.session.add(data)
#             db.session.commit()
#         return redirect("/")
#     return render_template("homepage.html")

# @app.route("/", methods=["POST", "GET"])
# def homepage():
#     db.create_all()
#     db.session.commit()
#     if request.method == "POST":
#         file = request.files['file']
#         path = f"static/{file.filename}"
#         file.save(path)
#         wb = openpyxl.load_workbook(path)
#         ws = wb.active
#         max_data = ws.max_row
#         for i in range(2,max_data+1):
#             sub_service = ws.cell(column=1, row=i).value
#             category = ws.cell(column=2, row=i).value
#             company = ws.cell(column=3, row=i).value
#             locality = ws.cell(column=4, row=i).value
#             address = ws.cell(column=5, row=i).value
#             pin = ws.cell(column=6, row=i).value
#             email_address = ws.cell(column=7, row=i).value
#             whatsapp = ws.cell(column=8, row=i).value
#             phone_1 = ws.cell(column=9, row=i).value
#             phone_2 = ws.cell(column=10, row=i).value
#             phone_3 = ws.cell(column=11, row=i).value
#             latitude = ws.cell(column=12, row=i).value
#             longitude = ws.cell(column=13, row=i).value
#             rating = ws.cell(column=14, row=i).value
#             reviews = ws.cell(column=15, row=i).value
#             website = ws.cell(column=16, row=i).value
#             verified= ws.cell(column=17, row=i).value
#             paid = ws.cell(column=18, row=i).value

#             user = User(sub_service=sub_service,category=category,company=company,locality=locality,address=address,pin=pin,email_address=email_address,whatsapp=whatsapp,phone_1=phone_1,phone_2=phone_2,phone_3=phone_3,latitude=latitude,longitude=longitude,rating=rating,reviews=reviews,website=website,verified=verified,paid=paid)
#             db.session.add(user)
#             db.session.commit()
#         return redirect("/")
#     return render_template("homepage.html")


if __name__ == "__main__":
    app.run(debug=True)
